"""Bulk master-data import for the Admin section.

Design goals (non-IT users):
- The client edits an Excel template they download, uploads it, sees a plain-language
  preview ("3 new, 1 updated, row 12 has an error"), and confirms. Nothing is written
  until the confirm step, and everything for one file is applied in one transaction.
- Deletion is never hard: records absent from the file are only deactivated when the
  admin ticks "Deactivate records missing from the file". History (SDMs, audit events)
  keeps its labels.

Supported lists and their file columns:
  employees   Name | Emp ID | Designation | Reports To Name | Reports To Emp ID
  lines       Name
  stations    Line Name | Station Name
  contractors Name

Employee rows become user logins (login = Emp ID, initial password 123) and their
supervisor links (users.superior_id) drive the automatic workflow routing.
"""
import csv
import io
import json
import re
import uuid
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from werkzeug.security import generate_password_hash

from . import db

STAGING_TTL = timedelta(hours=24)
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
STAGING_ID_PATTERN = re.compile(r"^[0-9a-f]{32}$")

ALLOWED_DESIGNATIONS = ["Station Controller", "Station Manager", "Line Manager", "Dy. HOD"]

# canonical key -> human column header (order shown in the templates)
KIND_HEADERS = {
    "employees": [
        ("name", "Name"),
        ("emp_id", "Emp ID"),
        ("designation", "Designation"),
        ("reports_to_name", "Reports To Name"),
        ("reports_to_emp_id", "Reports To Emp ID"),
    ],
    "lines": [("name", "Name")],
    "stations": [("line_name", "Line Name"), ("station_name", "Station Name")],
    "contractors": [("name", "Name")],
}

KIND_LABELS = {
    "employees": "Employees",
    "lines": "Lines",
    "stations": "Stations",
    "contractors": "Contractors",
}

KIND_TEMPLATE_FILENAME = {
    "employees": "employees-import-template.xlsx",
    "lines": "lines-import-template.xlsx",
    "stations": "stations-import-template.xlsx",
    "contractors": "contractors-import-template.xlsx",
}

# Roles that can appear in the employee file (protects admin/HOD/Cell logins).
EMPLOYEE_ROLES = set(db.DESIGNATION_ROLE_MAP.values())


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def staging_root():
    return Path(db.BASE_DIR) / "instance" / "import_staging"


def sweep_staging():
    root = staging_root()
    if not root.exists():
        return
    cutoff = datetime.now() - STAGING_TTL
    for path in root.iterdir():
        try:
            if path.is_file() and datetime.fromtimestamp(path.stat().st_mtime) < cutoff:
                path.unlink()
        except OSError:
            pass


def save_staging(kind, filename, content, deactivate_missing=False):
    """Persist an upload for the preview -> confirm two-step flow. Returns staging id."""
    sweep_staging()
    root = staging_root()
    root.mkdir(parents=True, exist_ok=True)
    staging_id = uuid.uuid4().hex
    ext = Path(filename).suffix.lower() or ".xlsx"
    (root / f"{staging_id}{ext}").write_bytes(content)
    meta = {
        "kind": kind,
        "ext": ext,
        "filename": Path(filename).name,
        "deactivate_missing": bool(deactivate_missing),
        "created": db.now_iso(),
    }
    (root / f"{staging_id}.json").write_text(json.dumps(meta), encoding="utf-8")
    return staging_id


def load_staging(staging_id):
    """Return (meta dict, file bytes) or None when the staging entry is gone/invalid."""
    if not STAGING_ID_PATTERN.match(staging_id or ""):
        return None
    root = staging_root()
    meta_path = root / f"{staging_id}.json"
    if not meta_path.exists():
        return None
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    file_path = root / f"{staging_id}{meta.get('ext', '.xlsx')}"
    if not file_path.exists():
        return None
    return meta, file_path.read_bytes()


def discard_staging(staging_id):
    if not STAGING_ID_PATTERN.match(staging_id or ""):
        return
    root = staging_root()
    for path in root.glob(f"{staging_id}.*"):
        try:
            path.unlink()
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Excel template download
# ---------------------------------------------------------------------------

def build_template_bytes(kind):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    headers = [label for _, label in KIND_HEADERS[kind]]
    sheet.append(headers)

    if kind == "employees":
        # Keep Emp ID as text so leading zeroes survive.
        sheet.cell(row=1, column=2).number_format = "@"
        # Designation dropdown for rows 2..2000.
        from openpyxl.worksheet.datavalidation import DataValidation

        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(ALLOWED_DESIGNATIONS) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Designation",
            error="Choose one of: " + ", ".join(ALLOWED_DESIGNATIONS),
        )
        sheet.add_data_validation(dv)
        dv.add(f"C2:C2000")

    # Instructions sheet.
    notes = workbook.create_sheet("Instructions")
    notes.append([f"How to update the {KIND_LABELS[kind].lower()} list"])
    notes.append([])
    notes.append(["1. Fill your data under the column header row on the 'Data' sheet."])
    notes.append(["2. Do not rename or delete the header row."])
    notes.append(["3. Upload the file on the Import page and review the preview before confirming."])
    notes.append(["4. Wrong rows are reported with their Excel row number - fix only those and upload again."])
    if kind == "employees":
        notes.append([])
        notes.append(["Column tips:"])
        notes.append(["  Name: full name of the employee."])
        notes.append(["  Emp ID: the employment number - it becomes the login. Keep numbers as text (no commas)."])
        notes.append(["  Designation: pick from the dropdown. Recognised: " + ", ".join(ALLOWED_DESIGNATIONS) + "."])
        notes.append(["  Reports To Name / Reports To Emp ID: the supervisor this employee works under."])
        notes.append(["    Give the Emp ID when you know it - it is more reliable than the name."])
        notes.append(["    The supervisor must also exist (as a row in this file or already in the system)."])
        notes.append([])
        notes.append(["What happens:"])
        notes.append(["  - New rows create logins (Emp ID) with the default password 123."])
        notes.append(["  - Changed names/designations update the account."])
        notes.append(["  - Missing rows are NOT touched unless you tick 'Deactivate records missing from the file'."])
    else:
        notes.append([])
        notes.append(["What happens:"])
        notes.append(["  - New rows are added."])
        notes.append(["  - Existing rows that were deactivated are reactivated."])
        notes.append(["  - Missing rows are NOT touched unless you tick 'Deactivate records missing from the file'."])
    if kind == "stations":
        notes.append([])
        notes.append(["Station Name: as shown in the system (it usually includes the station code, e.g. ARJANGARH (AJG))."])
        notes.append(["Line Name: must exactly match an existing Line (check the Lines list first)."])

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for column_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in column_cells), default=10)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def parse_file(kind, filename, content):
    """Return {'rows': [{'row_number': int, 'values': {key: str}}], 'error': str|None,
               'warnings': [str]} from raw uploaded bytes."""
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.worksheets[0]
            matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
        except Exception:
            return {"rows": [], "error": "Could not read the Excel file. Please re-save it as .xlsx from the downloaded template.", "warnings": []}
    elif ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            return {"rows": [], "error": "Could not read the CSV file. Save it as UTF-8 (Excel: File > Save As > CSV UTF-8) and try again.", "warnings": []}
        matrix = [list(row) for row in csv.reader(io.StringIO(text))]
    else:
        return {"rows": [], "error": "Unsupported file type. Please upload an .xlsx or .csv file.", "warnings": []}
    return matrix_to_rows(kind, matrix)


def matrix_to_rows(kind, matrix):
    warnings = []
    expected_labels = [label for _, label in KIND_HEADERS[kind]]

    non_empty = [idx for idx, row in enumerate(matrix) if any(clean(cell) for cell in row)]
    if not non_empty:
        return {"rows": [], "error": "The file is empty.", "warnings": warnings}

    header_row_idx = None
    column_by_label = {}
    for candidate in non_empty[:15]:
        row_values = [clean(cell) for cell in matrix[candidate]]
        if not row_values:
            continue
        found = {label.lower(): i for i, label in enumerate(row_values)}
        if all(label.lower() in found for label in expected_labels):
            header_row_idx = candidate
            column_by_label = found
            break

    if header_row_idx is None:
        wanted = ", ".join(expected_labels)
        return {
            "rows": [],
            "error": f"The file's first rows do not match the template. The columns must be: {wanted}. Please download the template and copy your data into it.",
            "warnings": warnings,
        }

    extra = [
        label
        for label in column_by_label
        if label not in {expected.lower() for expected in expected_labels}
    ]
    if extra:
        warnings.append("Ignored extra column(s): " + ", ".join(sorted(extra)[:3]).title())

    rows = []
    for idx in range(header_row_idx + 1, len(matrix)):
        raw = matrix[idx]
        if not any(clean(cell) for cell in raw):
            continue
        values = {}
        for key, label in KIND_HEADERS[kind]:
            column = column_by_label[label.lower()]
            values[key] = clean(raw[column]) if column < len(raw) else ""
        rows.append({"row_number": idx + 1, "values": values})

    if not rows:
        return {"rows": [], "error": "No data rows found below the header. Add rows in the template and upload again.", "warnings": warnings}
    return {"rows": rows, "error": None, "warnings": warnings}


# ---------------------------------------------------------------------------
# Planning (validation + classification, no writes)
# ---------------------------------------------------------------------------

def _plan_lines(plan, rows):
    seen = {}
    for entry in rows:
        row_number = entry["row_number"]
        name = entry["values"]["name"]
        if not name:
            _error(plan, row_number, "(empty)", "Name is required.")
            continue
        if name.lower() in seen:
            _error(plan, row_number, name, f'Duplicate Name "{name}" in the file (first seen on row {seen[name.lower()]}).')
            continue
        seen[name.lower()] = row_number

        existing = db.query_one("SELECT is_active FROM lines WHERE name = ?", (name,))
        if existing:
            if existing["is_active"]:
                _item(plan, row_number, "NO_CHANGE", name, "")
            else:
                _item(plan, row_number, "UPDATE", name, "Record exists but was deactivated. It will be reactivated.", {"name": name})
        else:
            _item(plan, row_number, "NEW", name, "Will be added.", {"name": name})


def _plan_stations(plan, rows):
    seen = set()
    for entry in rows:
        row_number = entry["row_number"]
        line_name = entry["values"]["line_name"]
        station_name = entry["values"]["station_name"]
        subject = f"{station_name} ({line_name})"
        if not station_name or not line_name:
            _error(plan, row_number, subject or "(empty)", "Both Line Name and Station Name are required.")
            continue
        key = (line_name.lower(), station_name.lower())
        if key in seen:
            _error(plan, row_number, subject, f'Duplicate station "{station_name}" under line "{line_name}" in the file.')
            continue
        seen.add(key)

        line = db.query_one("SELECT id, is_active FROM lines WHERE name = ?", (line_name,))
        if not line:
            _error(plan, row_number, subject, f'Line "{line_name}" was not found. Add the line first or check the spelling.')
            continue
        if not line["is_active"]:
            _error(plan, row_number, subject, f'Line "{line_name}" is deactivated. Reactivate the line first.')
            continue

        existing = db.query_one("SELECT is_active FROM stations WHERE line_id = ? AND name = ?", (line["id"], station_name))
        values = {"line_name": line_name, "station_name": station_name}
        if existing:
            if existing["is_active"]:
                _item(plan, row_number, "NO_CHANGE", subject, "")
            else:
                _item(plan, row_number, "UPDATE", subject, "Record exists but was deactivated. It will be reactivated.", values)
        else:
            _item(plan, row_number, "NEW", subject, "Will be added.", values)


def _plan_contractors(plan, rows):
    seen = {}
    for entry in rows:
        row_number = entry["row_number"]
        name = entry["values"]["name"]
        if not name:
            _error(plan, row_number, "(empty)", "Name is required.")
            continue
        if name.lower() in seen:
            _error(plan, row_number, name, f'Duplicate Name "{name}" in the file (first seen on row {seen[name.lower()]}).')
            continue
        seen[name.lower()] = row_number

        existing = db.query_one("SELECT is_active FROM contractors WHERE name = ?", (name,))
        if existing:
            if existing["is_active"]:
                _item(plan, row_number, "NO_CHANGE", name, "")
            else:
                _item(plan, row_number, "UPDATE", name, "Record exists but was deactivated. It will be reactivated.", {"name": name})
        else:
            _item(plan, row_number, "NEW", name, "Will be added.", {"name": name})


def _resolve_supervisor(values, file_emp_ids):
    """Resolve a row's Reports To reference to an Emp ID or a plain-language error.

    A supervisor may be someone already in the system or another row in this file.
    """
    emp_target = values.get("reports_to_emp_id") or ""
    name_target = values.get("reports_to_name") or ""
    if not emp_target and not name_target:
        return {"ok": True, "emp_id": None}

    if emp_target:
        if emp_target in file_emp_ids:
            return {"ok": True, "emp_id": emp_target}
        user = db.query_one("SELECT emp_id, is_active FROM users WHERE emp_id = ?", (emp_target,))
        if not user:
            return {"ok": False, "message": f'Reports To Emp ID "{emp_target}" was not found. Include the supervisor as a row in the file or remove the value.'}
        if not user["is_active"]:
            return {"ok": False, "message": f'Reports To Emp ID "{emp_target}" belongs to a deactivated account. Reactivate the supervisor first.'}
        return {"ok": True, "emp_id": emp_target}

    matches = db.query_all("SELECT emp_id FROM users WHERE name = ? AND is_active IS TRUE", (name_target,))
    if len(matches) > 1:
        return {"ok": False, "message": f'More than one active employee is named "{name_target}". Use the Reports To Emp ID column instead.'}
    if not matches:
        return {"ok": False, "message": f'Reports To "{name_target}" was not found. If the supervisor is new and only in this file, use the Reports To Emp ID column.'}
    return {"ok": True, "emp_id": matches[0]["emp_id"]}


def _plan_employees(plan, rows):
    seen = {}
    file_emp_ids = set()
    for entry in rows:
        emp_id = entry["values"]["emp_id"]
        if emp_id and emp_id not in seen:
            seen[emp_id] = entry["row_number"]
            file_emp_ids.add(emp_id)

    for entry in rows:
        row_number = entry["row_number"]
        values = entry["values"]
        name = values["name"]
        emp_id = values["emp_id"]
        designation = values["designation"]
        subject = f"{name or '(no name)'} ({emp_id or 'no Emp ID'})"

        if not name:
            _error(plan, row_number, subject, "Name is required.")
            continue
        if not emp_id:
            _error(plan, row_number, subject, "Emp ID is required.")
            continue
        if seen.get(emp_id) != row_number:
            _error(plan, row_number, subject, f'Duplicate Emp ID "{emp_id}" in the file (first seen on row {seen.get(emp_id)}).')
            continue
        if designation not in ALLOWED_DESIGNATIONS:
            allowed = ", ".join(ALLOWED_DESIGNATIONS)
            _error(plan, row_number, subject, f'Designation "{designation}" is not recognised. Use one of: {allowed}.')
            continue

        supervisor = _resolve_supervisor(values, file_emp_ids)
        if not supervisor["ok"]:
            _error(plan, row_number, subject, supervisor["message"])
            continue

        payload = {
            "name": name,
            "emp_id": emp_id,
            "designation": designation,
            "reports_to_emp_id": values.get("reports_to_emp_id") or "",
            "reports_to_name": values.get("reports_to_name") or "",
        }
        existing = db.query_one("SELECT name, designation, role, is_active FROM users WHERE emp_id = ?", (emp_id,))
        if existing:
            if existing["role"] not in EMPLOYEE_ROLES:
                _error(
                    plan,
                    row_number,
                    subject,
                    f'Emp ID "{emp_id}" belongs to an admin/HOD/Cell login ({existing["role"]}), which cannot be managed from the employee file. Remove the row or manage it on the Admin page.',
                )
                continue
            changed = existing["name"] != name or existing["designation"] != designation
            if existing["is_active"] and not changed:
                _item(plan, row_number, "NO_CHANGE", subject, "")
            else:
                detail = []
                if not existing["is_active"]:
                    detail.append("Account was deactivated; it will be reactivated.")
                if changed:
                    detail.append("Name/designation will be updated.")
                _item(plan, row_number, "UPDATE", subject, " ".join(detail), payload)
        else:
            _item(plan, row_number, "NEW", subject, "Will create a login (Emp ID) with the default password 123.", payload)


def _active_records_missing_from_file(kind, rows):
    """Existing active records not present in the file, for the optional deactivate step."""
    if kind == "employees":
        keys_in_file = {e["values"]["emp_id"] for e in rows if e["values"]["emp_id"]}
        return [
            {
                "action": "DEACTIVATE",
                "row_number": None,
                "subject": f"{u['name']} ({u['emp_id']})",
                "message": "Not in the file. The account login will be deactivated.",
                "values": {"emp_id": u["emp_id"]},
            }
            for u in db.query_all("SELECT name, emp_id, role FROM users WHERE is_active IS TRUE")
            if u["role"] in EMPLOYEE_ROLES and u["emp_id"] not in keys_in_file
        ]

    keys_in_file = set()
    for entry in rows:
        values = entry["values"]
        if kind == "lines" and values["name"]:
            keys_in_file.add(("n", values["name"].lower()))
        elif kind == "contractors" and values["name"]:
            keys_in_file.add(("n", values["name"].lower()))
        elif kind == "stations" and values["station_name"] and values["line_name"]:
            keys_in_file.add(("s", values["line_name"].lower(), values["station_name"].lower()))

    result = []
    if kind == "lines":
        for line in db.query_all("SELECT name FROM lines WHERE is_active IS TRUE"):
            if ("n", line["name"].lower()) not in keys_in_file:
                result.append(
                    {
                        "action": "DEACTIVATE",
                        "row_number": None,
                        "subject": line["name"],
                        "message": "Not in the file. The line and its stations will be deactivated.",
                        "values": {"name": line["name"]},
                    }
                )
    elif kind == "stations":
        for station in db.query_all(
            "SELECT st.name AS station_name, l.name AS line_name FROM stations st JOIN lines l ON l.id = st.line_id WHERE st.is_active IS TRUE"
        ):
            if ("s", station["line_name"].lower(), station["station_name"].lower()) not in keys_in_file:
                result.append(
                    {
                        "action": "DEACTIVATE",
                        "row_number": None,
                        "subject": f"{station['station_name']} ({station['line_name']})",
                        "message": "Not in the file. The station will be deactivated.",
                        "values": {"line_name": station["line_name"], "station_name": station["station_name"]},
                    }
                )
    elif kind == "contractors":
        for contractor in db.query_all("SELECT name FROM contractors WHERE is_active IS TRUE"):
            if ("n", contractor["name"].lower()) not in keys_in_file:
                result.append(
                    {
                        "action": "DEACTIVATE",
                        "row_number": None,
                        "subject": contractor["name"],
                        "message": "Not in the file. The contractor will be deactivated.",
                        "values": {"name": contractor["name"]},
                    }
                )
    return result


def _item(plan, row_number, action, subject, message, values=None):
    plan["items"].append(
        {"row_number": row_number, "action": action, "subject": subject, "message": message, "values": values or {}}
    )


def _error(plan, row_number, subject, message):
    _item(plan, row_number, "ERROR", subject, message)


def plan(kind, rows, deactivate_missing=False):
    planned = {"items": [], "kind": kind, "deactivate_missing": bool(deactivate_missing), "error": None}
    if kind == "employees":
        _plan_employees(planned, rows)
    elif kind == "lines":
        _plan_lines(planned, rows)
    elif kind == "stations":
        _plan_stations(planned, rows)
    elif kind == "contractors":
        _plan_contractors(planned, rows)
    else:
        planned["error"] = "Unknown list type."
        return planned

    if planned["deactivate_missing"]:
        planned["items"].extend(_active_records_missing_from_file(kind, rows))

    counts = {"NEW": 0, "UPDATE": 0, "DEACTIVATE": 0, "NO_CHANGE": 0, "ERROR": 0}
    for item in planned["items"]:
        counts[item["action"]] = counts.get(item["action"], 0) + 1
    planned["counts"] = counts
    planned["has_errors"] = counts["ERROR"] > 0
    return planned


# ---------------------------------------------------------------------------
# Applying (one transaction, re-plans first for safety)
# ---------------------------------------------------------------------------

def apply(kind, rows, deactivate_missing=False):
    planned = plan(kind, rows, deactivate_missing=deactivate_missing)
    if planned["has_errors"]:
        raise ValueError(f"{planned['counts']['ERROR']} row(s) have errors. Fix them and upload again.")

    database = db.get_db()
    try:
        if kind == "employees":
            _apply_employees(planned)
        else:
            _apply_master_lists(planned)
        database.commit()
    except Exception:
        database.rollback()
        raise
    return planned["counts"]


def _apply_master_lists(planned):
    database = db.get_db()
    for item in planned["items"]:
        action = item["action"]
        if action not in ("NEW", "UPDATE", "DEACTIVATE"):
            continue
        kind = planned["kind"]
        values = item["values"]
        if kind == "lines":
            name = values["name"]
            if action == "NEW":
                database.execute("INSERT INTO lines (name, is_active) VALUES (?, TRUE)", (name,))
            elif action == "UPDATE":
                database.execute("UPDATE lines SET is_active = TRUE WHERE name = ?", (name,))
            else:
                database.execute("UPDATE lines SET is_active = FALSE WHERE name = ?", (name,))
                database.execute(
                    "UPDATE stations SET is_active = FALSE WHERE line_id = (SELECT id FROM lines WHERE name = ?)",
                    (name,),
                )
        elif kind == "contractors":
            name = values["name"]
            if action == "NEW":
                database.execute("INSERT INTO contractors (name, is_active) VALUES (?, TRUE)", (name,))
            elif action == "UPDATE":
                database.execute("UPDATE contractors SET is_active = TRUE WHERE name = ?", (name,))
            else:
                database.execute("UPDATE contractors SET is_active = FALSE WHERE name = ?", (name,))
        elif kind == "stations":
            line_name = values["line_name"]
            station_name = values["station_name"]
            if action == "NEW":
                line = database.execute("SELECT id FROM lines WHERE name = ?", (line_name,)).fetchone()
                database.execute("INSERT INTO stations (line_id, name, is_active) VALUES (?, ?, TRUE)", (line["id"], station_name))
            elif action == "UPDATE":
                database.execute(
                    "UPDATE stations SET is_active = TRUE WHERE name = ? AND line_id = (SELECT id FROM lines WHERE name = ?)",
                    (station_name, line_name),
                )
            else:
                database.execute(
                    "UPDATE stations SET is_active = FALSE WHERE name = ? AND line_id = (SELECT id FROM lines WHERE name = ?)",
                    (station_name, line_name),
                )


def _apply_employees(planned):
    database = db.get_db()
    password_hash = generate_password_hash("123")

    pending = [item for item in planned["items"] if item["action"] in ("NEW", "UPDATE")]
    for item in pending:
        values = item["values"]
        role = db.DESIGNATION_ROLE_MAP.get(values["designation"], "STATION_CONTROLLER")
        if item["action"] == "NEW":
            database.execute(
                """
                INSERT INTO users (username, password_hash, name, emp_id, designation, role, is_admin, is_active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, FALSE, TRUE, ?)
                """,
                (values["emp_id"], password_hash, values["name"], values["emp_id"], values["designation"], role, db.now_iso()),
            )
        else:
            database.execute(
                """
                UPDATE users
                SET name = ?, designation = ?, role = ?, is_active = TRUE
                WHERE emp_id = ?
                """,
                (values["name"], values["designation"], role, values["emp_id"]),
            )

    # Second pass: supervisor + line assignment (needs all inserted rows visible).
    for item in pending:
        values = item["values"]
        supervisor = None
        if values.get("reports_to_emp_id"):
            supervisor = database.execute("SELECT id FROM users WHERE emp_id = ?", (values["reports_to_emp_id"],)).fetchone()
        elif values.get("reports_to_name"):
            supervisor = database.execute(
                "SELECT id FROM users WHERE name = ? AND is_active IS TRUE ORDER BY id LIMIT 1",
                (values["reports_to_name"],),
            ).fetchone()
        supervisor_id = supervisor["id"] if supervisor else None

        current = database.execute("SELECT line_id FROM users WHERE emp_id = ?", (values["emp_id"],)).fetchone()
        line_id = current["line_id"] if current and current["line_id"] else None
        if line_id is None and supervisor_id:
            sup = database.execute("SELECT line_id FROM users WHERE id = ?", (supervisor_id,)).fetchone()
            line_id = sup["line_id"] if sup else None
        database.execute(
            "UPDATE users SET superior_id = ?, line_id = COALESCE(line_id, ?) WHERE emp_id = ?",
            (supervisor_id, line_id, values["emp_id"]),
        )

    for item in planned["items"]:
        if item["action"] == "DEACTIVATE":
            database.execute("UPDATE users SET is_active = FALSE WHERE emp_id = ?", (item["values"]["emp_id"],))
